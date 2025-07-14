# Added comments explaining each step
import glob
import logging
import os
import time
import warnings

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Suppress unnecessary openpyxl warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Locate folders and files
coms_folder = r"C:\\Users\\perryr\\OneDrive - PIERER Mobility AG\\Desktop\\CF TEST"
forecast_file = os.path.join(coms_folder, "2025 NA Cash Forecast.xlsx")

# Look for multiple filename patterns: KTM Issued-* and KTM-Issued-*
coms_files = []
coms_files.extend(glob.glob(os.path.join(coms_folder, "KTM Issued-*.xlsx")))
coms_files.extend(glob.glob(os.path.join(coms_folder, "KTM-Issued-*.xlsx")))

# Find latest COMS file
if not coms_files:
    raise FileNotFoundError("No COMS file found.")
coms_file = max(coms_files, key=os.path.getmtime)
logger.info(f"Using COMS file: {os.path.basename(coms_file)}")

# Try loading workbook with retries for potential lock
for attempt in range(5):
    try:
        wb = load_workbook(forecast_file)
        break
    except PermissionError:
        time.sleep(2)
else:
    raise PermissionError("Cannot open forecast file after retries.")

# Read Week Table and build lookup dictionary
week_table = pd.read_excel(forecast_file, sheet_name='Week Table', header=1)
week_table['Day'] = pd.to_datetime(week_table['Day']).dt.normalize()
week_lookup = dict(zip(week_table['Day'], week_table['Week']))

# Load Details data from COMS file
details = pd.read_excel(coms_file, sheet_name='Details')

# Define payee groups and their target sheets
payee_mapping = {
    "COMS USD": ["61199", "184216", "23852", "226939", "163376", "184258", "252574", "252573", "239710", "239709",
                 "245753", "245755", "234679"],
    "COMS CAD": ["125593", "219929", "171291", "219930", "252647", "252568", "252558", "239760", "239712", "239761",
                 "239758", "245797", "245798", "236636"]
}

# Determine where to insert new sheets after COMS Summary
insert_idx = wb.sheetnames.index('COMS Summary') + 1 if 'COMS Summary' in wb.sheetnames else len(wb.sheetnames)

# Process each payee group
for sheet_name, payees in payee_mapping.items():
    filtered = details[details['Payee Nbr'].isin([int(x) for x in payees])].copy()

    # Use only Planned Issuance Date
    filtered['Issue Date'] = pd.to_datetime(filtered['Planned Issuance Date'], errors='coerce').dt.normalize()
    logger.info(f"{sheet_name}: using Planned Issuance Date for all Issue Dates.")

    filtered['Purchase Date'] = pd.to_datetime(filtered['Trans Date'], errors='coerce').dt.normalize()
    filtered['Purchase Month'] = filtered['Purchase Date'].dt.month

    # Lookup week numbers using Issue Date
    filtered['Week'] = filtered['Issue Date'].map(lambda x: week_lookup.get(x, ''))
    filtered['Week_Number'] = filtered['Week'].astype(str).str.replace('Week ', '', regex=False)

    # Build Current Forecast using Week, unless date not found then mark as PY
    filtered['Current Forecast'] = filtered['Payee Nbr'].astype(str) + 'Week ' + filtered['Week_Number']
    filtered.loc[~filtered['Issue Date'].isin(week_lookup.keys()), 'Current Forecast'] = \
        filtered.loc[~filtered['Issue Date'].isin(week_lookup.keys()), 'Payee Nbr'].astype(str) + 'PY'

    # Prepare final dataframe for writing
    final_df = filtered[
        ['Current Forecast', 'Issue Date', 'Inv/CM #', 'Net Amt', 'Week', 'Purchase Date', 'Purchase Month']].copy()
    final_df.columns = ['Current Forecast', 'Issue Date', 'Issue Number', 'Net Amount', 'Week', 'Purchase Date',
                        'Purchase Month']
    final_df['Comment'] = ''

    # Create or update worksheet
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name, insert_idx)
    insert_idx += 1

    # Load existing data from sheet to maintain YTD
    existing_data = [[c for c in row[:8]] for row in ws.iter_rows(min_row=3, values_only=True) if any(row)]
    existing_df = pd.DataFrame(existing_data,
                               columns=['Current Forecast', 'Issue Date', 'Issue Number', 'Net Amount', 'Week',
                                        'Purchase Date', 'Purchase Month', 'Comment']) if existing_data else pd.DataFrame()

    # Merge with new data, drop duplicates but keep newest
    combined_df = pd.concat([existing_df, final_df], ignore_index=True)
    combined_df.sort_values(by='Issue Date', ascending=False, inplace=True)
    combined_df = combined_df.drop_duplicates(
        subset=['Issue Number', 'Net Amount', 'Purchase Date'], keep='first'
    ).sort_values(by='Issue Date')

    # Rebuild the worksheet with header note
    ws.delete_rows(2, ws.max_row)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    cell = ws.cell(1, 1, value=" ")
    cell.alignment = Alignment(wrap_text=True, horizontal='center')
    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    cell.font = Font(color='FF0000', bold=True)

    # Write headers and data
    for col_idx, col_name in enumerate(combined_df.columns, start=1):
        ws.cell(2, col_idx, col_name)
    for row_idx, row in enumerate(combined_df.itertuples(index=False), start=3):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row_idx, col_idx, value)

    # Auto-size columns based on content
    for col_idx in range(1, 9):
        letter = get_column_letter(col_idx)
        max_len = max((len(str(cell.value)) if cell.value else 0 for cell in ws[letter]), default=0)
        ws.column_dimensions[letter].width = min(max_len + 2, 50)

# Save the updated workbook
wb.save(forecast_file)
logger.info("Process completed. Updated sheets.")
